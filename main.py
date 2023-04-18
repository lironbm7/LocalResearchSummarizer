import subprocess
import re
import json
import csv
import string
import os
import yaml
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# You don't have to touch anything in main.py for the program to work
# Define config.yaml according to Readme / predefined config examples

def main():

    # Extract definitions from config.yaml
    with open('config.yaml', 'r') as f:
        config = yaml.safe_load(f)
    directories = config['directories']
    output_file = config['output_file']
    tool_version = config['tool_version']
    tool_cmd = config['tool_cmd']
    headers = config['headers']
    jq_cmd = config['jq_cmd']
    tool_cmd = config['tool_cmd']
    tool_cmd = [arg.replace('{output_file}', output_file) for arg in tool_cmd]
    

    # Auto-defined variables
    tool_name = tool_cmd[0]
    # if tool version is undefined, try to get it automatically. if '--version' does not work, don't mention the version.
    if tool_version == 'undefined':
        try:
            tool_version = f'{subprocess.check_output([tool_name, "--version"]).decode().strip()}'
            output_file_name = f"{tool_name}-{tool_version}"  # use the automatically generated version
        except subprocess.CalledProcessError:
            output_file_name = f"{tool_name}"
    else:
        output_file_name = f"{tool_name}-{tool_version}"  # use the hard-coded version


    # Run CLI commands for each one of the directories defined above
    for directory in directories:

        # clean up (delete old findings files) from a previous scan to prevent false positives and only maintain new data
        cleanup_files = [output_file, "parsed.json", "parsed.csv"]
        for file_name in cleanup_files:
            file_path = os.path.join(directory, file_name)
            if os.path.exists(file_path):
                os.remove(file_path)

        # Extract project name from directory path
        project_name = re.search(r'/([^/]+)/?$', directory).group(1)
        print(f'Current working directory: {project_name}')
        
        # Change working directory
        subprocess.run(["cd", directory])
        
        # Run tool scan
        subprocess.run(tool_cmd, cwd=directory)
        
        # Parse tool output using jq
        if os.path.exists(f'{directory}/{output_file}') and os.path.getsize(f'{directory}/{output_file}') > 0:
            with open(f"{directory}/parsed.json", "w") as f:
                subprocess.run(jq_cmd, cwd=directory, stdout=f, input=open(f"{directory}/{output_file}").read().encode())
        
        # Convert parsed JSON files to CSV files
        if os.path.exists(f'{directory}/parsed.json') and os.path.getsize(f'{directory}/parsed.json') > 0:
            with open(f"{directory}/parsed.json") as f:
                data = json.load(f)
            with open(f"{directory}/parsed.csv", "w", newline='') as f:
                writer = csv.writer(f)
                headers_row = ['Repository'] + headers
                writer.writerow(headers_row)
                for item in data:
                    # Handle non-printable ASCII characters in conversion from .json to .csv
                    data_row = [project_name] + ["".join(filter(lambda x: x in string.printable, str(item[header]))) for header in headers]
                    writer.writerow(data_row)
        else:
            print(f'> {project_name} has no findings therefore no parsed.csv. Skipping..')

    # Merge findings from all directories into a single CSV file
    with open(f'{output_file_name}.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Repository'] + headers)
        for directory in directories:
            project_name = re.search(r'/([^/]+)/?$', directory).group(1)
            if os.path.isfile(f'{directory}/parsed.csv'):
                with open(f"{directory}/parsed.csv") as subdir_csv:
                    subdir_reader = csv.reader(subdir_csv)
                    next(subdir_reader)  # skip header row
                    for row in subdir_reader:
                        # Write all fields except the Repository field
                        writer.writerow([project_name] + row[1:])
    
    # convert .csv to .xlsx
    with open(f'{output_file_name}.csv', 'r') as f:
        data = csv.reader(f)
        next(data)  # Skip header row

        # Create a new Excel workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        ws.append(["Repository"] + headers)

        # Apply bold and fill pattern to header row
        header_row = ws[1]
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        for cell in header_row:
            cell.font = header_font
            cell.fill = header_fill

        # Define fill for alternating rows
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        grey_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

        # Loop through the rows of data and write them to the worksheet
        for i, row in enumerate(data, start=2):  # Start at row 2 to skip header
            ws.append(row)

            # Apply alternating fill pattern to data rows only
            if i % 2 == 0:
                for cell in ws[f'A{i}':f'Z{i}'][0]:
                    cell.fill = white_fill
            else:
                for cell in ws[f'A{i}':f'Z{i}'][0]:
                    cell.fill = grey_fill

        # Auto-size columns to fit their content
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Save the Excel workbook
        wb.save(f'{output_file_name}.xlsx')
        print(f'> Findings saved into an Excel sheet named {output_file_name}.xlsx')

    # summarize findings in .txt
    with open(f'{output_file_name}.csv', 'r') as f:
        data = csv.reader(f)
        next(data)  # Skip header row
        # Create a .txt file to summarize what the program did
        repositories = [re.search(r'/([^/]+)/?$', directory).group(1) for directory in directories]
        total_repositories_scanned = len(repositories)
        timestamp = datetime.now().strftime("%d-%m-%Y")
        lines_in_csv = sum(1 for _ in data)  # count instances in csv for the summarization .txt file later on

        # Create .txt file and write the required information
        with open(f'{output_file_name}.txt', 'w') as t:
            t.write(f"{tool_name} - {tool_version}\n")
            t.write(f"{timestamp}\n\n")
            t.write("CLI COMMANDS USED:\n")
            t.write(f"{' '.join(tool_cmd)}\n")
            t.write(f"{' '.join(jq_cmd)}\n\n")
            t.write(f"Total Repositories Scanned: {total_repositories_scanned}\n")
            t.write("Repositories: " + ', '.join(repositories) + "\n")
            t.write(f"Total Findings: {lines_in_csv}\n")
            print(f'> Report Sumumarization saved into a Text file named {output_file_name}.txt')

if __name__ == "__main__":
    csv.field_size_limit(10485760) # 10 MB
    main()
